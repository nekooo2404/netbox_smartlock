import csv
import json
import tempfile
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
import openpyxl

from core.models import ObjectType
from dcim.models import Device, DeviceRole, DeviceType, Location, Manufacturer, Rack, Region, Site
from netbox.choices import CSVDelimiterChoices, ImportFormatChoices
from users.models import ObjectPermission

from upload_file_plugin.models import UploadedFile
from upload_file_plugin.services import sync_uploaded_files

from netbox_smartlock.api.serializers import SmartLockSerializer
from netbox_smartlock.contracts import SMARTLOCK_IMPORT_FIELD_NAMES
from netbox_smartlock.models import Asset, AssetGroup, SmartLock


class SmartLockIntegrationTestBase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="smartlock-user",
            password="test-password",
        )
        cls.region = Region.objects.create(name="Region 1", slug="region-1")
        cls.site = Site.objects.create(name="Site 1", slug="site-1", region=cls.region)
        cls.location = Location.objects.create(name="Room A", slug="room-a", site=cls.site)
        cls.rack = Rack.objects.create(name="R01", site=cls.site, location=cls.location)
        cls.manufacturer = Manufacturer.objects.create(name="GTSC", slug="gtsc")
        cls.device_type = DeviceType.objects.create(
            manufacturer=cls.manufacturer,
            model="DCIM Asset",
            slug="dcim-asset",
        )
        cls.device_role = DeviceRole.objects.create(name="Asset", slug="asset", color="2196f3")
        cls.asset_group = AssetGroup.objects.create(
            name="Door Locks",
            slug="door-locks",
            code="DL",
            status=AssetGroup.STATUS_ACTIVE,
        )

    @staticmethod
    def csv_text(rows):
        return "\n".join(rows)

    @classmethod
    def grant_object_permission(cls, model, actions):
        permission = ObjectPermission(name=f"{model.__name__} {','.join(actions)}", actions=actions)
        permission.save()
        permission.users.add(cls.user)
        permission.object_types.add(ObjectType.objects.get_for_model(model))
        return permission

    @staticmethod
    def grant_object_permission_to_user(user, model, actions, name_prefix="test", constraints=None):
        permission = ObjectPermission(
            name=f"{name_prefix} {user.username} {model._meta.label_lower} {','.join(actions)}",
            actions=actions,
            constraints=constraints,
        )
        permission.save()
        permission.users.add(user)
        permission.object_types.add(ObjectType.objects.get_for_model(model))
        return permission

    @classmethod
    def grant_smartlock_import_permissions(cls):
        cls.grant_object_permission(SmartLock, ["add", "change"])
        for model in (AssetGroup, Region, Site, Location, Rack):
            cls.grant_object_permission(model, ["view"])

    def login(self):
        self.client.force_login(self.user)

    def make_device_asset(self, *, name="Shared Device Asset", asset_group=None):
        asset_group = asset_group or self.asset_group
        device = Device.objects.create(
            name=name,
            asset_tag=f"AT-{name.replace(' ', '-').upper()[:20]}",
            device_type=self.device_type,
            role=self.device_role,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )
        return Asset.objects.create(
            name=name,
            code=device.asset_tag,
            device=device,
            asset_group=asset_group,
            status=Asset.STATUS_ACTIVE,
        )

    def smartlock_row(self, *, code="SL-001", name="Front Door", rack_lookup="site-1|room-a|R01"):
        return (
            f",{name},{code},door-locks,active,,,"
            f"Smart Lock,M1,S1,GTSC,2026-01-01,2026-01-10,12,"
            f"region-1,site-1,room-a,{rack_lookup},front"
        )


class SmartLockCoreImportTest(SmartLockIntegrationTestBase):
    def post_import(self, rows):
        self.login()
        self.grant_smartlock_import_permissions()
        return self.client.post(
            reverse("plugins:netbox_smartlock:smartlock_bulk_import"),
            {
                "format": ImportFormatChoices.CSV,
                "csv_delimiter": CSVDelimiterChoices.COMMA,
                "data": self.csv_text(rows),
            },
        )

    def test_core_bulk_import_creates_smartlock_with_rack_mapping(self):
        response = self.post_import(
            [
                ",".join(SMARTLOCK_IMPORT_FIELD_NAMES),
                self.smartlock_row(),
            ]
        )

        self.assertEqual(response.status_code, 302)
        smartlock = SmartLock.objects.get(code="SL-001")
        self.assertEqual(smartlock.rack, self.rack)
        self.assertEqual(smartlock.site, self.site)
        self.assertEqual(smartlock.location, self.location)
        self.assertEqual(smartlock.region, self.region)
        self.assertEqual(smartlock.warranty_expiration_date.isoformat(), "2027-01-10")

    def test_core_bulk_import_is_atomic_when_later_row_is_invalid(self):
        response = self.post_import(
            [
                ",".join(SMARTLOCK_IMPORT_FIELD_NAMES),
                self.smartlock_row(code="SL-VALID"),
                self.smartlock_row(code="SL-BAD", rack_lookup="missing-rack"),
            ]
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(SmartLock.objects.filter(code__in=["SL-VALID", "SL-BAD"]).exists())

    def test_core_csv_export_round_trips_to_core_import_contract(self):
        smartlock = SmartLock.objects.create(
            name="Back Door",
            code="SL-EXPORT",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_ACTIVE,
            device_type="Smart Lock",
            model="M2",
            serial="S2",
            manufacturer="GTSC",
            bought_date="2026-01-10",
            warranty_period=12,
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
            rack_face=SmartLock.RACK_FACE_FRONT,
        )
        self.login()
        self.grant_object_permission(SmartLock, ["view"])

        response = self.client.get(reverse("plugins:netbox_smartlock:smartlock_list"), {"export": "table"})

        self.assertEqual(response.status_code, 200)
        rows = list(csv.DictReader(response.content.decode().splitlines()))
        row = next(item for item in rows if item["code"] == smartlock.code)
        self.assertEqual(tuple(row.keys()), SMARTLOCK_IMPORT_FIELD_NAMES)
        self.assertEqual(row["asset_group"], self.asset_group.slug)
        self.assertEqual(row["rack_lookup"], "site-1|room-a|R01")


class AssetGroupCoreImportTest(SmartLockIntegrationTestBase):
    def post_import(self, rows):
        self.login()
        self.grant_object_permission(AssetGroup, ["add", "change"])
        return self.client.post(
            reverse("plugins:netbox_smartlock:assetgroup_bulk_import"),
            {
                "format": ImportFormatChoices.CSV,
                "csv_delimiter": CSVDelimiterChoices.COMMA,
                "data": self.csv_text(rows),
            },
        )

    def test_assetgroup_core_import_url_exists(self):
        self.login()
        self.grant_object_permission(AssetGroup, ["add"])

        response = self.client.get(reverse("plugins:netbox_smartlock:assetgroup_bulk_import"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Import Nhóm tài sản")

    def test_assetgroup_list_does_not_render_missing_import_url(self):
        self.login()
        self.grant_object_permission(AssetGroup, ["view", "add"])

        response = self.client.get(reverse("plugins:netbox_smartlock:assetgroup_list"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertNotIn("/plugins/smartlock/asset-groups/None", content)
        self.assertIn("/plugins/smartlock/asset-groups/import/", content)

    def test_assetgroup_core_import_creates_group(self):
        response = self.post_import(
            [
                "name,slug,code,status,exclude_from_visualization,description,comments",
                "Cabinet Locks,cabinet-locks,CL,active,true,Imported group,",
            ]
        )

        self.assertEqual(response.status_code, 302)
        asset_group = AssetGroup.objects.get(slug="cabinet-locks")
        self.assertEqual(asset_group.code, "CL")
        self.assertTrue(asset_group.exclude_from_visualization)
        self.assertEqual(asset_group.description, "Imported group")

    def test_assetgroup_import_requires_code(self):
        response = self.post_import(
            [
                "name,slug,code,status,description,comments",
                "Missing Code,missing-code,,active,Invalid group,",
            ]
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(AssetGroup.objects.filter(slug="missing-code").exists())

    def test_assetgroup_import_rejects_description_longer_than_dcim_limit(self):
        response = self.post_import(
            [
                "name,slug,code,status,description,comments",
                f"Long Description,long-description,LD,active,{ 'x' * 501 },",
            ]
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(AssetGroup.objects.filter(slug="long-description").exists())

    def test_assetgroup_list_defaults_to_dcim_spec_columns(self):
        self.login()
        self.grant_object_permission(AssetGroup, ["view", "add"])

        response = self.client.get(reverse("plugins:netbox_smartlock:assetgroup_list"))

        self.assertEqual(response.status_code, 200)
        for label in ("Tên", "Mã", "Trạng thái", "Mô tả", "Người tạo", "Thời gian tạo", "Thời gian cập nhật"):
            self.assertContains(response, label)

    def test_assetgroup_visualization_label_is_vietnamese_in_ui(self):
        self.login()
        self.grant_object_permission(AssetGroup, ["view", "add"])

        list_response = self.client.get(
            reverse("plugins:netbox_smartlock:assetgroup_list"),
            {"columns": "name,exclude_from_visualization"},
        )
        add_response = self.client.get(reverse("plugins:netbox_smartlock:assetgroup_add"))
        detail_response = self.client.get(self.asset_group.get_absolute_url())

        for response in (list_response, add_response, detail_response):
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Loại trừ khỏi visualization")

    def test_assetgroup_quick_search_matches_vietnamese_status_label(self):
        inactive_group = AssetGroup.objects.create(
            name="Retired Locks",
            slug="retired-locks",
            code="RL",
            status=AssetGroup.STATUS_INACTIVE,
        )
        self.login()
        self.grant_object_permission(AssetGroup, ["view"])

        response = self.client.get(reverse("plugins:netbox_smartlock:assetgroup_list"), {"q": "Không hoạt động"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, inactive_group.name)
        self.assertNotContains(response, self.asset_group.name)
        self.assertContains(response, "Không hoạt động")
        self.assertNotIn(">Inactive<", response.content.decode())

    def test_assetgroup_form_renders_cancel_confirmation_modal(self):
        self.login()
        self.grant_object_permission(AssetGroup, ["add"])

        response = self.client.get(reverse("plugins:netbox_smartlock:assetgroup_add"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-bs-target="#smartlock-cancel-confirm-modal"')
        self.assertContains(response, "Hủy thay đổi?")

    def test_assetgroup_excel_export_uses_xlsx_content_type(self):
        self.login()
        self.grant_object_permission(AssetGroup, ["view"])

        response = self.client.get(
            reverse("plugins:netbox_smartlock:assetgroup_list"),
            {"assetgroup_export": "excel_report"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("netbox_asset_groups.xlsx", response["Content-Disposition"])

    def test_asset_model_reuses_dcim_device_without_copying_infrastructure(self):
        asset = self.make_device_asset(name="Mapped Device Asset")

        self.assertEqual(asset.device.site, self.site)
        self.assertEqual(asset.device.location, self.location)
        self.assertEqual(asset.device.rack, self.rack)
        self.assertEqual(asset.device_type, self.device_type)
        self.assertEqual(asset.manufacturer, self.manufacturer)

    def test_device_asset_warranty_expiration_is_calculated_on_asset_model(self):
        asset = self.make_device_asset(name="Warranty Device")
        asset.bought_date = "2026-01-31"
        asset.warranty_period = 1
        asset.save()

        asset.refresh_from_db()

        self.assertEqual(asset.warranty_expiration_date.isoformat(), "2026-02-28")

    def test_device_asset_defaults_status_on_asset_model(self):
        asset = self.make_device_asset(name="Default Status Asset")

        self.assertEqual(asset.status, Asset.STATUS_ACTIVE)

    def test_assetgroup_detail_lists_dcim_assets_using_shared_group(self):
        asset = self.make_device_asset(name="Door Controller 01")
        self.login()
        self.grant_object_permission(AssetGroup, ["view"])
        self.grant_object_permission(Asset, ["view"])

        response = self.client.get(self.asset_group.get_absolute_url())

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tài sản")
        self.assertContains(response, asset.name)
        self.assertContains(response, f"asset_group_id={self.asset_group.pk}")

    def test_device_asset_list_uses_dcim_device_with_dicm_default_columns(self):
        asset = self.make_device_asset(name="List Device")
        self.login()
        self.grant_object_permission(Asset, ["view", "add", "change", "delete"])

        response = self.client.get(reverse("plugins:netbox_smartlock:device_asset_list"))

        self.assertEqual(response.status_code, 200)
        for label in (
            "Tên", "Mã", "Trạng thái", "Thiết bị", "Địa điểm", "Vị trí", "Nhóm tài sản",
            "Hãng sản xuất", "Loại thiết bị", "Người tạo", "Thời gian tạo", "Thời gian cập nhật",
        ):
            self.assertContains(response, label)
        self.assertNotContains(response, "Thiết bị DCIM")
        self.assertContains(response, asset.name)
        self.assertContains(response, asset.code)
        self.assertContains(response, reverse("plugins:netbox_smartlock:device_asset_add"))
        self.assertContains(response, reverse("plugins:netbox_smartlock:device_asset_edit", kwargs={"pk": asset.pk}))
        self.assertContains(response, reverse("plugins:netbox_smartlock:device_asset_delete", kwargs={"pk": asset.pk}))

    def test_device_asset_quick_search_matches_asset_group_and_status(self):
        active_asset = self.make_device_asset(name="Active Group Asset")
        backup_asset = self.make_device_asset(name="Backup Group Asset")
        backup_asset.status = Asset.STATUS_BACKUP
        backup_asset.save()
        self.login()
        self.grant_object_permission(Asset, ["view"])

        group_response = self.client.get(
            reverse("plugins:netbox_smartlock:device_asset_list"),
            {"q": self.asset_group.name},
        )
        status_response = self.client.get(
            reverse("plugins:netbox_smartlock:device_asset_list"),
            {"q": "Dự phòng"},
        )

        self.assertEqual(group_response.status_code, 200)
        self.assertContains(group_response, active_asset.name)
        self.assertContains(group_response, backup_asset.name)
        self.assertEqual(status_response.status_code, 200)
        self.assertContains(status_response, backup_asset.name)
        self.assertNotContains(status_response, active_asset.name)

    def test_device_asset_add_requires_asset_code(self):
        self.login()
        self.grant_object_permission(Asset, ["view", "add"])
        self.grant_object_permission(Device, ["view"])
        self.grant_object_permission(AssetGroup, ["view"])
        device = Device.objects.create(
            name="Missing Code Device",
            device_type=self.device_type,
            role=self.device_role,
            site=self.site,
            location=self.location,
        )

        response = self.client.post(
            reverse("plugins:netbox_smartlock:device_asset_add"),
            {
                "name": "Missing Asset Code",
                "code": "",
                "device": device.pk,
                "asset_group": self.asset_group.pk,
                "status": Asset.STATUS_ACTIVE,
                "upload_files": "[]",
                "_create": "Tạo",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFormError(response.context["form"], "code", "This field is required.")
        self.assertFalse(Asset.objects.filter(name="Missing Asset Code").exists())

    def test_device_asset_form_renders_cancel_confirmation_modal(self):
        self.login()
        self.grant_object_permission(Asset, ["add"])
        self.grant_object_permission(Device, ["view"])
        self.grant_object_permission(AssetGroup, ["view"])

        response = self.client.get(reverse("plugins:netbox_smartlock:device_asset_add"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Thiết bị")
        self.assertNotContains(response, "Thiết bị DCIM")
        self.assertContains(response, 'data-bs-target="#smartlock-cancel-confirm-modal"')
        self.assertContains(response, "Hủy thay đổi?")

    def test_device_asset_detail_labels_device_without_dcim_suffix(self):
        asset = self.make_device_asset(name="Detail Label Asset")
        self.login()
        self.grant_object_permission(Asset, ["view"])

        response = self.client.get(asset.get_absolute_url())

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Thiết bị")
        self.assertNotContains(response, "Thiết bị DCIM")

    def test_device_asset_add_redirects_to_asset_list_after_create(self):
        self.login()
        self.grant_object_permission(Asset, ["view", "add"])
        self.grant_object_permission(Device, ["view"])
        self.grant_object_permission(AssetGroup, ["view"])
        device = Device.objects.create(
            name="Created Device",
            asset_tag="ASSET-CREATED-DEVICE",
            device_type=self.device_type,
            role=self.device_role,
            site=self.site,
            location=self.location,
        )

        response = self.client.post(
            reverse("plugins:netbox_smartlock:device_asset_add"),
            {
                "name": "Created Asset",
                "code": "ASSET-CREATED",
                "device": device.pk,
                "asset_group": self.asset_group.pk,
                "status": Asset.STATUS_ACTIVE,
                "upload_files": "[]",
                "_create": "Tạo",
            },
        )

        self.assertRedirects(response, reverse("plugins:netbox_smartlock:device_asset_list"))
        self.assertTrue(Asset.objects.filter(name="Created Asset", code="ASSET-CREATED", device=device).exists())

    def test_device_asset_list_export_uses_scoped_device_table(self):
        asset = self.make_device_asset(name="Export Asset")
        self.login()
        self.grant_object_permission(Asset, ["view"])

        response = self.client.get(
            reverse("plugins:netbox_smartlock:device_asset_list"),
            {"export": "table"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, asset.name)
        self.assertContains(response, asset.code)

    def test_device_asset_excel_export_uses_xlsx_content_type(self):
        asset = self.make_device_asset(name="Excel Asset")
        self.login()
        self.grant_object_permission(Asset, ["view"])

        response = self.client.get(
            reverse("plugins:netbox_smartlock:device_asset_list"),
            {"device_asset_export": "excel_report"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("netbox_assets.xlsx", response["Content-Disposition"])

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        flattened_cells = [str(cell.value) for row in worksheet.iter_rows() for cell in row if cell.value is not None]
        self.assertIn(asset.name, flattened_cells)
        self.assertIn(asset.code, flattened_cells)

    def test_device_asset_file_view_renders_upload_widget(self):
        asset = self.make_device_asset(name="File Asset")
        self.login()
        self.grant_object_permission(Asset, ["view", "change"])

        response = self.client.get(
            reverse("plugins:netbox_smartlock:device_asset_files", kwargs={"pk": asset.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "File đính kèm tài sản")
        self.assertContains(response, 'data-model-name="asset"')

    def test_device_asset_file_view_requires_change_permission_for_post(self):
        asset = self.make_device_asset(name="File Permission Asset")
        self.login()
        self.grant_object_permission(Asset, ["view"])

        response = self.client.post(
            reverse("plugins:netbox_smartlock:device_asset_files", kwargs={"pk": asset.pk}),
            {
                "upload_files": "[]",
                "_update": "Lưu",
            },
        )

        self.assertEqual(response.status_code, 403)

    def test_device_detail_includes_asset_file_button_and_panel(self):
        asset = self.make_device_asset(name="Detail File Asset")
        self.login()
        self.grant_object_permission(Asset, ["view", "change"])
        self.grant_object_permission(Device, ["view"])

        response = self.client.get(asset.device.get_absolute_url())

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "File tài sản")
        self.assertContains(response, "File đính kèm tài sản")

    def test_device_detail_hides_asset_files_without_asset_view_permission(self):
        asset = self.make_device_asset(name="Hidden File Asset")
        self.login()
        self.grant_object_permission(Device, ["view"])

        response = self.client.get(asset.device.get_absolute_url())

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "File tài sản")
        self.assertNotContains(response, "File đính kèm tài sản")

    def test_assetgroup_api_exposes_visualization_flag(self):
        self.asset_group.exclude_from_visualization = True
        self.asset_group.save()
        self.login()
        self.grant_object_permission(AssetGroup, ["view"])

        response = self.client.get(
            reverse("plugins-api:netbox_smartlock-api:assetgroup-detail", kwargs={"pk": self.asset_group.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["exclude_from_visualization"])

    def test_asset_api_exposes_asset_model_with_nested_dcim_device(self):
        asset = self.make_device_asset(name="API Asset")
        self.login()
        self.grant_object_permission(Asset, ["view"])

        response = self.client.get(
            reverse("plugins-api:netbox_smartlock-api:asset-detail", kwargs={"pk": asset.pk})
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["name"], asset.name)
        self.assertEqual(payload["code"], asset.code)
        self.assertEqual(payload["device"]["id"], asset.device_id)
        self.assertEqual(payload["asset_group"]["id"], self.asset_group.pk)


class SmartLockCrudViewTest(SmartLockIntegrationTestBase):
    def test_smartlock_list_defaults_to_dcim_spec_metadata_columns(self):
        SmartLock.objects.create(
            name="List Lock",
            code="SL-LIST",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_ACTIVE,
            device_type="Smart Lock",
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )
        self.login()
        self.grant_object_permission(SmartLock, ["view"])

        response = self.client.get(reverse("plugins:netbox_smartlock:smartlock_list"))

        self.assertEqual(response.status_code, 200)
        for label in ("Người tạo", "Thời gian tạo", "Thời gian cập nhật"):
            self.assertContains(response, label)

    def test_smartlock_quick_search_matches_asset_group_and_vietnamese_status_label(self):
        active_lock = SmartLock.objects.create(
            name="Active Search Lock",
            code="SL-ACTIVE-SEARCH",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_ACTIVE,
            device_type="Smart Lock",
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )
        backup_lock = SmartLock.objects.create(
            name="Backup Search Lock",
            code="SL-BACKUP-SEARCH",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_BACKUP,
            device_type="Smart Lock",
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )
        self.login()
        self.grant_object_permission(SmartLock, ["view"])

        group_response = self.client.get(
            reverse("plugins:netbox_smartlock:smartlock_list"),
            {"q": self.asset_group.name},
        )
        status_response = self.client.get(
            reverse("plugins:netbox_smartlock:smartlock_list"),
            {"q": "Dự phòng"},
        )

        self.assertEqual(group_response.status_code, 200)
        self.assertContains(group_response, active_lock.name)
        self.assertContains(group_response, backup_lock.name)
        self.assertEqual(status_response.status_code, 200)
        self.assertContains(status_response, backup_lock.name)
        self.assertNotContains(status_response, active_lock.name)
        self.assertContains(status_response, "Dự phòng")
        self.assertNotIn(">Backup<", status_response.content.decode())

    def test_add_view_creates_smartlock_without_server_error(self):
        self.login()
        self.grant_object_permission(SmartLock, ["add"])
        for model in (AssetGroup, Region, Site, Location, Rack):
            self.grant_object_permission(model, ["view"])

        response = self.client.post(
            reverse("plugins:netbox_smartlock:smartlock_add"),
            {
                "name": "Front Door UI",
                "code": "SL-UI-001",
                "asset_group": self.asset_group.pk,
                "status": SmartLock.STATUS_ACTIVE,
                "device_type": "Smart Lock",
                "model": "",
                "serial": "",
                "manufacturer": "",
                "setup_date": "",
                "bought_date": "",
                "warranty_period": "",
                "region": self.region.pk,
                "site": self.site.pk,
                "location": self.location.pk,
                "rack": self.rack.pk,
                "rack_face": "",
                "description": "",
                "comments": "",
                "upload_files": "[]",
            },
        )

        self.assertEqual(response.status_code, 302)
        smartlock = SmartLock.objects.get(code="SL-UI-001")
        self.assertEqual(smartlock.site, self.site)
        self.assertEqual(smartlock.location, self.location)
        self.assertEqual(smartlock.rack, self.rack)

    def test_smartlock_and_assetgroup_changelog_pages_render(self):
        smartlock = SmartLock.objects.create(
            name="Changelog Lock",
            code="SL-CHANGELOG",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_ACTIVE,
            device_type="Smart Lock",
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )
        self.login()
        self.grant_object_permission(SmartLock, ["view"])
        self.grant_object_permission(AssetGroup, ["view"])

        smartlock_response = self.client.get(
            reverse("plugins:netbox_smartlock:smartlock_changelog", kwargs={"pk": smartlock.pk})
        )
        assetgroup_response = self.client.get(
            reverse("plugins:netbox_smartlock:assetgroup_changelog", kwargs={"pk": self.asset_group.pk})
        )

        self.assertEqual(smartlock_response.status_code, 200, smartlock_response.content.decode())
        self.assertEqual(assetgroup_response.status_code, 200, assetgroup_response.content.decode())

    def test_smartlock_form_renders_cancel_confirmation_modal(self):
        self.login()
        self.grant_object_permission(SmartLock, ["add"])
        for model in (AssetGroup, Region, Site, Location, Rack):
            self.grant_object_permission(model, ["view"])

        response = self.client.get(reverse("plugins:netbox_smartlock:smartlock_add"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-bs-target="#smartlock-cancel-confirm-modal"')
        self.assertContains(response, "Hủy thay đổi?")

    def test_detail_renders_warranty_state_as_netbox_badge(self):
        smartlock = SmartLock.objects.create(
            name="Warranty Detail Lock",
            code="SL-WARRANTY",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_ACTIVE,
            device_type="Smart Lock",
            bought_date="2026-01-10",
            warranty_period=24,
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )
        self.login()
        self.grant_object_permission(SmartLock, ["view"])

        response = self.client.get(smartlock.get_absolute_url())

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            '<span class="badge text-bg-success">Còn bảo hành</span>',
            html=True,
        )
        self.assertNotContains(response, "&lt;span class=&quot;badge text-bg-success&quot;&gt;")

    def test_smartlock_excel_export_strips_warranty_badge_html(self):
        smartlock = SmartLock.objects.create(
            name="Warranty Excel Lock",
            code="SL-WARRANTY-XLSX",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_ACTIVE,
            device_type="Smart Lock",
            bought_date="2026-01-10",
            warranty_period=24,
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )
        self.login()
        self.grant_object_permission(SmartLock, ["view"])

        response = self.client.get(
            reverse("plugins:netbox_smartlock:smartlock_list"),
            {"smartlock_export": "excel_report"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("netbox_smartlocks.xlsx", response["Content-Disposition"])

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        flattened_cells = [str(cell.value) for row in worksheet.iter_rows() for cell in row if cell.value is not None]
        self.assertIn(smartlock.name, flattened_cells)
        self.assertIn("Còn bảo hành", flattened_cells)
        self.assertFalse(any("<span" in value for value in flattened_cells))

    def test_add_form_rejects_out_of_scope_dcim_related_objects(self):
        other_region = Region.objects.create(name="Restricted Region UI", slug="restricted-region-ui")
        other_site = Site.objects.create(name="Restricted Site UI", slug="restricted-site-ui", region=other_region)
        other_location = Location.objects.create(name="Restricted Room UI", slug="restricted-room-ui", site=other_site)
        other_rack = Rack.objects.create(name="RR01", site=other_site, location=other_location)

        self.login()
        self.grant_object_permission(SmartLock, ["add"])
        self.grant_object_permission(AssetGroup, ["view"])
        self.grant_object_permission_to_user(
            self.user,
            Region,
            ["view"],
            name_prefix="smartlock-form-scope",
            constraints={"slug": self.region.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Site,
            ["view"],
            name_prefix="smartlock-form-scope",
            constraints={"slug": self.site.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Location,
            ["view"],
            name_prefix="smartlock-form-scope",
            constraints={"slug": self.location.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Rack,
            ["view"],
            name_prefix="smartlock-form-scope",
            constraints={"name": self.rack.name},
        )

        response = self.client.post(
            reverse("plugins:netbox_smartlock:smartlock_add"),
            {
                "name": "Out Of Scope UI",
                "code": "SL-UI-SCOPE-BAD",
                "asset_group": self.asset_group.pk,
                "status": SmartLock.STATUS_ACTIVE,
                "device_type": "Smart Lock",
                "model": "",
                "serial": "",
                "manufacturer": "",
                "setup_date": "",
                "bought_date": "",
                "warranty_period": "",
                "region": other_region.pk,
                "site": other_site.pk,
                "location": other_location.pk,
                "rack": other_rack.pk,
                "rack_face": SmartLock.RACK_FACE_FRONT,
                "description": "",
                "comments": "",
                "upload_files": "[]",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(SmartLock.objects.filter(code="SL-UI-SCOPE-BAD").exists())


class SmartLockPermissionTest(SmartLockIntegrationTestBase):
    def test_core_import_view_requires_add_permission(self):
        self.login()
        response = self.client.get(reverse("plugins:netbox_smartlock:smartlock_bulk_import"))

        self.assertEqual(response.status_code, 403)

    def test_core_import_view_allows_netbox_object_permission(self):
        self.login()
        self.grant_object_permission(SmartLock, ["add"])
        response = self.client.get(reverse("plugins:netbox_smartlock:smartlock_bulk_import"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Import Khóa thông minh")

    def test_core_import_rejects_out_of_scope_dcim_rack_lookup(self):
        other_region = Region.objects.create(name="Restricted Region Import", slug="restricted-region-import")
        other_site = Site.objects.create(name="Restricted Site Import", slug="restricted-site-import", region=other_region)
        other_location = Location.objects.create(name="Restricted Room Import", slug="restricted-room-import", site=other_site)
        Rack.objects.create(name="IR01", site=other_site, location=other_location)

        self.login()
        self.grant_object_permission(SmartLock, ["add"])
        self.grant_object_permission(AssetGroup, ["view"])
        self.grant_object_permission_to_user(
            self.user,
            Region,
            ["view"],
            name_prefix="smartlock-import-scope",
            constraints={"slug": self.region.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Site,
            ["view"],
            name_prefix="smartlock-import-scope",
            constraints={"slug": self.site.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Location,
            ["view"],
            name_prefix="smartlock-import-scope",
            constraints={"slug": self.location.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Rack,
            ["view"],
            name_prefix="smartlock-import-scope",
            constraints={"name": self.rack.name},
        )

        response = self.client.post(
            reverse("plugins:netbox_smartlock:smartlock_bulk_import"),
            {
                "format": ImportFormatChoices.CSV,
                "csv_delimiter": CSVDelimiterChoices.COMMA,
                "data": self.csv_text(
                    [
                        ",".join(SMARTLOCK_IMPORT_FIELD_NAMES),
                        self.smartlock_row(
                            code="SL-IMPORT-SCOPE-BAD",
                            name="Out Of Scope Import",
                            rack_lookup="restricted-site-import|restricted-room-import|IR01",
                        ).replace("region-1,site-1,room-a,", ",,,"),
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(SmartLock.objects.filter(code="SL-IMPORT-SCOPE-BAD").exists())


class SmartLockApiValidationTest(SmartLockIntegrationTestBase):
    def test_api_serializer_uses_shared_rack_mapping_validation(self):
        data = {
            "name": "  API Lock  ",
            "code": "  SL-API  ",
            "status": SmartLock.STATUS_ACTIVE,
            "asset_group_id": self.asset_group.pk,
            "device_type": "  Smart Lock  ",
            "region_id": self.region.pk,
            "site_id": self.site.pk,
            "location_id": self.location.pk,
            "rack_id": self.rack.pk,
            "rack_face": SmartLock.RACK_FACE_FRONT,
        }

        serializer = SmartLockSerializer(data=data)

        self.assertTrue(serializer.is_valid(), serializer.errors)
        self.assertEqual(serializer.validated_data["name"], "API Lock")
        self.assertEqual(serializer.validated_data["code"], "SL-API")
        self.assertEqual(serializer.validated_data["device_type"], "Smart Lock")
        self.assertEqual(serializer.validated_data["site"], self.site)
        self.assertEqual(serializer.validated_data["location"], self.location)
        self.assertEqual(serializer.validated_data["region"], self.region)

    def test_api_serializer_rejects_rack_site_mismatch(self):
        other_site = Site.objects.create(name="Site 2", slug="site-2", region=self.region)
        other_location = Location.objects.create(name="Room B", slug="room-b", site=other_site)
        data = {
            "name": "API Lock",
            "code": "SL-API-BAD",
            "status": SmartLock.STATUS_ACTIVE,
            "asset_group_id": self.asset_group.pk,
            "device_type": "Smart Lock",
            "region_id": self.region.pk,
            "site_id": other_site.pk,
            "location_id": other_location.pk,
            "rack_id": self.rack.pk,
        }

        serializer = SmartLockSerializer(data=data)

        self.assertFalse(serializer.is_valid())
        self.assertIn("rack", serializer.errors)

    def test_api_rejects_out_of_scope_dcim_related_objects(self):
        other_region = Region.objects.create(name="Restricted Region API", slug="restricted-region-api")
        other_site = Site.objects.create(name="Restricted Site API", slug="restricted-site-api", region=other_region)
        other_location = Location.objects.create(name="Restricted Room API", slug="restricted-room-api", site=other_site)
        other_rack = Rack.objects.create(name="AR01", site=other_site, location=other_location)

        self.login()
        self.grant_object_permission(SmartLock, ["add"])
        self.grant_object_permission(AssetGroup, ["view"])
        self.grant_object_permission_to_user(
            self.user,
            Region,
            ["view"],
            name_prefix="smartlock-api-scope",
            constraints={"slug": self.region.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Site,
            ["view"],
            name_prefix="smartlock-api-scope",
            constraints={"slug": self.site.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Location,
            ["view"],
            name_prefix="smartlock-api-scope",
            constraints={"slug": self.location.slug},
        )
        self.grant_object_permission_to_user(
            self.user,
            Rack,
            ["view"],
            name_prefix="smartlock-api-scope",
            constraints={"name": self.rack.name},
        )

        response = self.client.post(
            reverse("plugins-api:netbox_smartlock-api:smartlock-list"),
            data=json.dumps(
                {
                    "name": "Out Of Scope API",
                    "code": "SL-API-SCOPE-BAD",
                    "status": SmartLock.STATUS_ACTIVE,
                    "asset_group_id": self.asset_group.pk,
                    "device_type": "Smart Lock",
                    "region_id": other_region.pk,
                    "site_id": other_site.pk,
                    "location_id": other_location.pk,
                    "rack_id": other_rack.pk,
                    "rack_face": SmartLock.RACK_FACE_FRONT,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(SmartLock.objects.filter(code="SL-API-SCOPE-BAD").exists())

    def test_api_rejects_inactive_asset_group(self):
        inactive_asset_group = AssetGroup.objects.create(
            name="Inactive Locks",
            slug="inactive-locks",
            code="IL",
            status=AssetGroup.STATUS_INACTIVE,
        )

        self.login()
        self.grant_object_permission(SmartLock, ["add"])
        self.grant_object_permission(AssetGroup, ["view"])
        for model in (Region, Site, Location, Rack):
            self.grant_object_permission(model, ["view"])

        response = self.client.post(
            reverse("plugins-api:netbox_smartlock-api:smartlock-list"),
            data=json.dumps(
                {
                    "name": "Inactive Asset Group API",
                    "code": "SL-API-INACTIVE-GROUP",
                    "status": SmartLock.STATUS_ACTIVE,
                    "asset_group_id": inactive_asset_group.pk,
                    "device_type": "Smart Lock",
                    "region_id": self.region.pk,
                    "site_id": self.site.pk,
                    "location_id": self.location.pk,
                    "rack_id": self.rack.pk,
                    "rack_face": SmartLock.RACK_FACE_FRONT,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(SmartLock.objects.filter(code="SL-API-INACTIVE-GROUP").exists())


class SmartLockAttachmentLifecycleTest(SmartLockIntegrationTestBase):
    def make_smartlock(self):
        return SmartLock.objects.create(
            name="Attachment Lock",
            code="SL-FILE",
            asset_group=self.asset_group,
            status=SmartLock.STATUS_ACTIVE,
            device_type="Smart Lock",
            region=self.region,
            site=self.site,
            location=self.location,
            rack=self.rack,
        )

    def pending_payload(self, media_root, filename="lock.png", content=b"image-data"):
        tmp_dir = Path(media_root) / "uploads" / "tmp"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        tmp_file = tmp_dir / filename
        tmp_file.write_bytes(content)
        return [
            {
                "file_name": filename,
                "path": f"{settings.MEDIA_URL.rstrip('/')}/uploads/tmp/{filename}",
                "size": len(content),
            }
        ]

    def test_attachment_sync_add_remove_and_model_delete_cleanup(self):
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                smartlock = self.make_smartlock()

                sync_uploaded_files(
                    smartlock,
                    json.dumps(self.pending_payload(media_root)),
                    model_name="smartlock",
                )

                uploaded = UploadedFile.objects.get(model_name="smartlock", object_id=smartlock.pk)
                stored_path = Path(media_root) / uploaded.file.name
                self.assertTrue(stored_path.exists())

                sync_uploaded_files(smartlock, "[]", model_name="smartlock")
                self.assertFalse(UploadedFile.objects.filter(pk=uploaded.pk).exists())
                self.assertFalse(stored_path.exists())

                sync_uploaded_files(
                    smartlock,
                    json.dumps(self.pending_payload(media_root, filename="lock-2.png")),
                    model_name="smartlock",
                )
                uploaded = UploadedFile.objects.get(model_name="smartlock", object_id=smartlock.pk)
                stored_path = Path(media_root) / uploaded.file.name

                smartlock.delete()

                self.assertFalse(UploadedFile.objects.filter(pk=uploaded.pk).exists())
                self.assertFalse(stored_path.exists())
